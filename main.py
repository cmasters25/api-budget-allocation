from dotenv import load_dotenv
import os
load_dotenv()

from fastapi import FastAPI, UploadFile, File, Form
from litellm import completion, completion_cost
import traceback
from pypdf import PdfReader
import io
from docx import Document
from openpyxl import load_workbook
import psycopg2
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles

app = FastAPI()

#for ui
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

app.mount("/static", StaticFiles(directory="static"), name="static")

#connecting to postgresql to keep track of budgets
conn = psycopg2.connect(
    os.getenv("DATABASE_URL")
)

cursor = conn.cursor()

cursor.execute("""
CREATE TABLE IF NOT EXISTS teams (
    name TEXT PRIMARY KEY,
    budget DOUBLE PRECISION,
    spent DOUBLE PRECISION
)
""")

conn.commit()


#tracking budget
def get_spent(team: str) -> float:
    cursor.execute("SELECT spent FROM teams WHERE name=%s", (team,))
    row = cursor.fetchone()
    return row[0] if row else 0.0


def update_spent(team: str, new_spent: float):
    cursor.execute(
    "UPDATE teams SET spent=%s WHERE name=%s",
    (new_spent, team)
)
    conn.commit()

def get_budget(team: str):
    cursor.execute(
        "SELECT budget FROM teams WHERE name=%s",
        (team,)
    )

    row = cursor.fetchone()

    return row[0] if row else None

#budgets with postgresql
teams = [
    ("team 1", 10.0, 0.0),
    ("team 2", 10.0, 0.0),
    ("team 3", 10.0, 0.0),
    ("team 4", 10.0, 0.0),
    ("team 5", 10.0, 0.0),
    ("team 6", 10.0, 0.0),
    ("team 7", 10.0, 0.0),
    ("team 8", 10.0, 0.0),
    ("team 9", 10.0, 0.0),
]

for team in teams:
    cursor.execute(
        """
        INSERT INTO teams(name,budget,spent)
        VALUES (%s,%s,%s)
        ON CONFLICT (name) DO NOTHING
        """,
        team
    )

conn.commit()

#file upload info, supports .txt, word, excel, and pdf
async def extract_text(file: UploadFile):

    if file is None:
        return ""

    contents = await file.read()

    filename = file.filename.lower()

    # txt
    if filename.endswith(".txt"):
        return contents.decode("utf-8")

    # pdf
    elif filename.endswith(".pdf"):

        pdf = PdfReader(io.BytesIO(contents))

        text = ""

        for page in pdf.pages:
            page_text = page.extract_text()

            if page_text:
                text += page_text + "\n"

        return text

    # word
    elif filename.endswith(".docx"):

        doc = Document(io.BytesIO(contents))

        text = ""

        for paragraph in doc.paragraphs:
            text += paragraph.text + "\n"

        return text

    # excel
    elif filename.endswith(".xlsx"):

        workbook = load_workbook(io.BytesIO(contents))

        text = ""

        for sheet in workbook.worksheets:

            text += f"Worksheet: {sheet.title}\n"

            for row in sheet.iter_rows(values_only=True):

                row_text = " | ".join(
                    str(cell) if cell is not None else ""
                    for cell in row
                )

                text += row_text + "\n"

        return text

    else:
        raise ValueError("Unsupported file type.")
    
#homepage
@app.get("/")
async def home():
    return FileResponse("static/index.html")

#chat function
@app.post("/chat")
async def chat(
    team: str = Form(...),
    provider: str = Form(...),
    model: str = Form(...),
    message: str = Form(...),
    file: UploadFile = File(None)
):

    budget = get_budget(team)

    if budget is None:
        return {"error": "Unknown team"}

    spent = get_spent(team)

    remaining = budget - spent  

    if remaining <= 0:
        return {
            "status": "blocked",
            "error": "budget reached",
            "remaining_budget": 0
        }

    try:
        file_text = await extract_text(file)
    except Exception as e:
        return {
            "error": str(e)
        }
    
    full_prompt = f"""
    Student Prompt:
    {message}

    Uploaded File:
    {file_text}
    """

    try:
        response = completion(
            model=f"{provider}/{model}",
            messages=[
                {"role": "user", "content": full_prompt}
            ]
        )

        cost = completion_cost(completion_response=response)

        new_spent = spent + cost
        update_spent(team, new_spent)

        return {
            "status": "allowed",
            "response": response.choices[0].message.content,
            "team": team,
            "provider": provider,
            "model": model,
            "cost": cost,
            "spent": get_spent(team),
            "remaining": budget - new_spent
        }

    except Exception as e:
        return {"error": str(e)}